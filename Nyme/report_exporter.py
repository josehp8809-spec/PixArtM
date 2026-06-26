"""
report_exporter.py — Genera reportes en Excel (.xlsx) y PDF (.pdf).
"""
import io
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def export_excel(df, agent_stats, sentiments, start_dt, end_dt):
    """
    Genera un Excel con múltiples pestañas.
    Retorna bytes para st.download_button.
    """
    wb = Workbook()

    # Colores de marca
    COLOR_HEADER  = "0A84FF"
    COLOR_RED     = "FF3B30"
    COLOR_YELLOW  = "FF9F0A"
    COLOR_GREEN   = "30D158"
    COLOR_DARK    = "1C1C1E"

    def _style_header(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    period_str = f"{start_dt.strftime('%d/%m/%Y %H:%M')} — {end_dt.strftime('%d/%m/%Y %H:%M')}"

    # ── Pestaña 1: Resumen ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "📊 Resumen"
    ws1["A1"] = "NYME — Reporte de Actividad"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A2"] = f"Período: {period_str}"
    ws1["A3"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    ws1["A5"] = "Métrica"
    ws1["B5"] = "Valor"
    _style_header(ws1, 5, 2)

    total_msgs  = len(df)
    total_chats = df["wa_id"].nunique() if not df.empty else 0
    inbound     = len(df[df["type"] == "INBOUND"]) if not df.empty else 0
    outbound    = len(df[df["type"] != "INBOUND"]) if not df.empty else 0

    summary_rows = [
        ("Total mensajes", total_msgs),
        ("Conversaciones únicas", total_chats),
        ("Mensajes entrantes (clientes)", inbound),
        ("Mensajes salientes (agentes)", outbound),
    ]
    for i, (label, val) in enumerate(summary_rows, start=6):
        ws1[f"A{i}"] = label
        ws1[f"B{i}"] = val

    _auto_width(ws1)

    # ── Pestaña 2: Por Agente ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("👤 Por Agente")
    headers2 = ["Agente", "Enviados", "Recibidos", "Total", "Estado"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header(ws2, 1, len(headers2))

    # Construir stats por agente desde df
    agent_summary = {}
    if not df.empty:
        for _, row in df.iterrows():
            agent = row["agent_username"] or "Sin asignar"
            if agent not in agent_summary:
                agent_summary[agent] = {"sent": 0, "received": 0}
            if row["type"] == "INBOUND":
                agent_summary[agent]["received"] += 1
            else:
                agent_summary[agent]["sent"] += 1

    # Mapa de sentimientos por agente (aproximado — basado en conversaciones)
    sentiment_by_contact = {s[0]: s[1] for s in sentiments} if sentiments else {}

    for r, (agent, stats) in enumerate(agent_summary.items(), start=2):
        sent     = stats["sent"]
        received = stats["received"]
        total    = sent + received
        ws2.cell(row=r, column=1, value=f"@{agent}")
        ws2.cell(row=r, column=2, value=sent)
        ws2.cell(row=r, column=3, value=received)
        ws2.cell(row=r, column=4, value=total)
        ws2.cell(row=r, column=5, value="Ver sentimientos →")

    _auto_width(ws2)

    # ── Pestaña 3: Conversaciones ─────────────────────────────────────────────
    ws3 = wb.create_sheet("💬 Conversaciones")
    headers3 = ["Contacto", "Tipo", "Mensaje", "Agente", "Fecha/Hora"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    _style_header(ws3, 1, len(headers3))

    if not df.empty:
        for r, row in enumerate(df.itertuples(index=False), start=2):
            ws3.cell(row=r, column=1, value=f"+{row.wa_id}")
            ws3.cell(row=r, column=2, value=row.type)
            ws3.cell(row=r, column=3, value=str(row.body or "")[:200])
            ws3.cell(row=r, column=4, value=f"@{row.agent_username}" if row.agent_username else "—")
            ts = row.created_at
            ws3.cell(row=r, column=5, value=ts.strftime("%d/%m/%Y %H:%M") if ts else "")

    _auto_width(ws3)

    # ── Pestaña 4: Sentimientos ───────────────────────────────────────────────
    ws4 = wb.create_sheet("🧠 Sentimientos")
    headers4 = ["Contacto", "Sentimiento", "Urgente", "Razón detectada", "Analizado"]
    for col, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=col, value=h)
    _style_header(ws4, 1, len(headers4))

    SENT_COLORS = {"positivo": COLOR_GREEN, "negativo": COLOR_RED, "neutral": COLOR_YELLOW}
    for r, (wa_id, sentiment, urgent, reason, analyzed_at) in enumerate(sentiments or [], start=2):
        ws4.cell(row=r, column=1, value=f"+{wa_id}")
        cell_sent = ws4.cell(row=r, column=2, value=sentiment or "—")
        color = SENT_COLORS.get(sentiment, "CCCCCC")
        cell_sent.fill = PatternFill("solid", fgColor=color)
        ws4.cell(row=r, column=3, value="⚠️ Sí" if urgent else "No")
        ws4.cell(row=r, column=4, value=reason or "—")
        ws4.cell(row=r, column=5, value=analyzed_at.strftime("%d/%m/%Y %H:%M") if analyzed_at else "")

    _auto_width(ws4)

    # Serializar a bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# PDF EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def export_pdf(df, agent_stats, sentiments, start_dt, end_dt, agent_feedbacks=None):
    """
    Genera un PDF ejecutivo del reporte.
    Retorna bytes para st.download_button.
    """
    try:
        from fpdf import FPDF
    except ImportError:
        raise ImportError("Instala fpdf2: pip install fpdf2")

    period_str  = f"{start_dt.strftime('%d/%m/%Y %H:%M')} al {end_dt.strftime('%d/%m/%Y %H:%M')}"
    total_msgs  = len(df)
    total_chats = df["wa_id"].nunique() if not df.empty else 0
    inbound     = len(df[df["type"] == "INBOUND"]) if not df.empty else 0
    outbound    = len(df[df["type"] != "INBOUND"]) if not df.empty else 0

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # ── Portada ───────────────────────────────────────────────────────────────
    pdf.set_fill_color(0, 0, 0)
    pdf.rect(0, 0, 210, 297, "F")

    pdf.set_font("Helvetica", "B", 28)
    pdf.set_text_color(255, 255, 255)
    pdf.ln(30)
    pdf.cell(0, 12, "NYME", align="C", new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "", 14)
    pdf.set_text_color(180, 180, 180)
    pdf.cell(0, 8, "Reporte de Actividad — WhatsApp Platform", align="C", new_x="LMARGIN", new_y="NEXT")

    pdf.ln(10)
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(100, 150, 255)
    pdf.cell(0, 6, f"Período: {period_str}", align="C", new_x="LMARGIN", new_y="NEXT")

    pdf.ln(5)
    pdf.set_text_color(120, 120, 120)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}", align="C",
             new_x="LMARGIN", new_y="NEXT")

    # ── Métricas Globales ─────────────────────────────────────────────────────
    pdf.add_page()
    pdf.set_fill_color(28, 28, 30)
    pdf.rect(0, 0, 210, 297, "F")

    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(10, 132, 255)
    pdf.ln(10)
    pdf.cell(0, 10, "Métricas Globales del Período", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    metrics = [
        ("Total de mensajes", str(total_msgs)),
        ("Conversaciones únicas", str(total_chats)),
        ("Mensajes de clientes", str(inbound)),
        ("Mensajes de agentes", str(outbound)),
    ]
    for label, val in metrics:
        pdf.set_font("Helvetica", "", 12)
        pdf.set_text_color(200, 200, 200)
        pdf.cell(100, 9, f"  {label}:", new_x="RIGHT", new_y="TOP")
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(60, 9, val, new_x="LMARGIN", new_y="NEXT")

    # ── Análisis de Sentimientos ──────────────────────────────────────────────
    if sentiments:
        pdf.ln(8)
        pdf.set_font("Helvetica", "B", 16)
        pdf.set_text_color(10, 132, 255)
        pdf.cell(0, 10, "Análisis de Sentimientos", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

        pos = sum(1 for s in sentiments if s[1] == "positivo")
        neu = sum(1 for s in sentiments if s[1] == "neutral")
        neg = sum(1 for s in sentiments if s[1] == "negativo")
        urg = sum(1 for s in sentiments if s[2])

        for label, val, color in [
            ("Positivos 🟢", str(pos), (48, 209, 88)),
            ("Neutrales 🟡", str(neu), (255, 159, 10)),
            ("Negativos 🔴", str(neg), (255, 59, 48)),
            ("Urgentes ⚠️",  str(urg), (255, 59, 48)),
        ]:
            pdf.set_font("Helvetica", "", 12)
            pdf.set_text_color(200, 200, 200)
            pdf.cell(100, 9, f"  {label}:", new_x="RIGHT", new_y="TOP")
            pdf.set_font("Helvetica", "B", 12)
            pdf.set_text_color(*color)
            pdf.cell(60, 9, val, new_x="LMARGIN", new_y="NEXT")

        # Conversaciones urgentes
        urgent_list = [(s[0], s[3]) for s in sentiments if s[2]]
        if urgent_list:
            pdf.ln(5)
            pdf.set_font("Helvetica", "B", 13)
            pdf.set_text_color(255, 59, 48)
            pdf.cell(0, 9, "⚠️ Conversaciones que requieren atención:", new_x="LMARGIN", new_y="NEXT")
            for wa_id, reason in urgent_list[:10]:
                pdf.set_font("Helvetica", "", 11)
                pdf.set_text_color(220, 220, 220)
                pdf.cell(0, 7, f"  • +{wa_id}: {reason or 'Alerta detectada'}", new_x="LMARGIN", new_y="NEXT")

    # ── Feedback por Agente ───────────────────────────────────────────────────
    if agent_feedbacks:
        pdf.ln(8)
        pdf.set_font("Helvetica", "B", 16)
        pdf.set_text_color(10, 132, 255)
        pdf.cell(0, 10, "Feedback de Desempeño por Agente", new_x="LMARGIN", new_y="NEXT")

        for agent, feedback_text in agent_feedbacks.items():
            pdf.ln(5)
            pdf.set_font("Helvetica", "B", 13)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(0, 8, f"@{agent}", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(180, 180, 180)
            pdf.multi_cell(0, 6, feedback_text or "Sin datos suficientes para análisis.")

    # Serializar a bytes
    buf = io.BytesIO()
    buf.write(pdf.output())
    buf.seek(0)
    return buf.getvalue()
